import io
import json
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Query, Request, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import redis.asyncio as aioredis

from config import settings
from database import get_db
from models.models import Business, SearchLog, ViewLog, User, PointLog
from routers.auth import get_current_user

router = APIRouter(prefix="/search", tags=["search"])


# ── 마스킹 함수 (서버사이드 - 개발자도구로 볼 수 없음) ──────────────
def partial_name(name: str) -> str:
    """홍길동치킨 → 홍길●● (첫 2글자 공개)"""
    if not name:
        return "●●●"
    if len(name) <= 2:
        return name[0] + "●"
    return name[:2] + "●" * (len(name) - 2)


def partial_addr(addr: str | None) -> str:
    """서울시 강남구 테헤란로 → 시+구+로 까지만 공개, 이후 ●●● 처리"""
    if not addr:
        return "●●●"
    parts = addr.strip().split()
    if len(parts) <= 3:
        return " ".join(parts) + " ●●●"
    visible = " ".join(parts[:3])
    return visible + " ●●●"


def mask_tel(tel: str | None) -> str:
    """전화번호 마스킹"""
    if not tel or not tel.strip():
        return ""
    return "●●●-●●●●-●●●●"


def get_point_cost(b: Business) -> int:
    """전화번호 없으면 5P, 있으면 30P"""
    if not b.tel or not b.tel.strip():
        return 5
    return settings.POINT_PER_VIEW


def normalize_status(status: str | None) -> str:
    """영업/정상 → 영업중 정규화"""
    if not status:
        return ""
    if "영업" in status:
        return "영업중"
    return status


def build_business_item(b: Business, revealed: bool, rank: int, free: bool = False) -> dict:
    has_tel = bool(b.tel and b.tel.strip())
    point_cost = get_point_cost(b)
    status = normalize_status(b.status)

    if revealed or free:
        return {
            "id": str(b.id),
            "rank": rank,
            "bsn_nm": b.bsn_nm,
            "uptae_nm": b.uptae_nm,
            "addr": b.addr,
            "road_addr": b.road_addr,
            "tel": b.tel,
            "status": status,
            "open_date": str(b.open_date) if b.open_date else None,
            "sido": b.sido,
            "sigungu": b.sigungu,
            "masked": False,
            "has_tel": has_tel,
            "point_cost": point_cost,
        }
    return {
        "id": str(b.id),
        "rank": rank,
        "bsn_nm": partial_name(b.bsn_nm),
        "uptae_nm": b.uptae_nm,
        "addr": partial_addr(b.addr),
        "road_addr": partial_addr(b.road_addr),
        "tel": mask_tel(b.tel),
        "status": status,
        "open_date": str(b.open_date) if b.open_date else None,
        "sido": b.sido,
        "sigungu": b.sigungu,
        "masked": True,
        "has_tel": has_tel,
        "point_cost": point_cost,
    }


# ── 통계 ──────────────────────────────────────────────────────
@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_db)):
    """메인화면용 데이터 통계 (Redis 10분 캐시)"""
    CACHE_KEY = "stats:main"
    CACHE_TTL = 600  # 10분

    # Redis 캐시 확인
    r = aioredis.from_url(settings.REDIS_URL, decode_responses=True)
    try:
        cached = await r.get(CACHE_KEY)
        if cached:
            await r.aclose()
            return json.loads(cached)
    except Exception:
        pass

    # 캐시 없으면 DB 조회
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    monday_start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    total = (await db.execute(select(func.count()).select_from(Business))).scalar()
    this_month = (await db.execute(
        select(func.count()).select_from(Business).where(Business.created_at >= month_start)
    )).scalar()
    this_week = (await db.execute(
        select(func.count()).select_from(Business).where(Business.created_at >= monday_start)
    )).scalar()
    today_count = (await db.execute(
        select(func.count()).select_from(Business).where(Business.created_at >= today_start)
    )).scalar()

    result = {
        "total": total,
        "this_month": this_month,
        "this_week": this_week,
        "today": today_count,
        "updated_at": now.strftime("%Y-%m-%d"),
    }

    # Redis에 저장
    try:
        await r.setex(CACHE_KEY, CACHE_TTL, json.dumps(result))
    except Exception:
        pass
    finally:
        await r.aclose()

    return result


# ── 검색 ──────────────────────────────────────────────────────
@router.get("")
async def search(
    q: str = Query(..., min_length=1),
    page: int = Query(1, ge=1),
    size: int = Query(100, ge=1, le=1000),
    sido: str | None = None,
    sigungu: str | None = None,
    uptae: str | None = None,
    status: str | None = None,
    request: Request = None,
    db: AsyncSession = Depends(get_db),
):
    # 현재 사용자 확인 (옵셔널)
    token = request.headers.get("Authorization", "").replace("Bearer ", "") if request else None
    current_user: User | None = None
    viewed_ids: set[str] = set()

    if token:
        try:
            from jose import jwt
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
            user_id = payload.get("sub")
            if user_id:
                import uuid
                result = await db.execute(select(User).where(User.id == uuid.UUID(user_id)))
                current_user = result.scalar_one_or_none()
                vl = await db.execute(
                    select(ViewLog.business_id).where(ViewLog.user_id == current_user.id)
                )
                viewed_ids = {str(row[0]) for row in vl.fetchall()}
        except Exception:
            pass

    # ── 동의어 사전 ────────────────────────────────────────────────
    SYNONYMS = {
        "미용실": "미용업", "헤어샵": "미용업", "헤어": "미용업", "뷰티": "미용업",
        "수영장": "수영장업",
        "헬스장": "체력단련장", "헬스": "체력단련장", "피트니스": "체력단련장", "gym": "체력단련장",
        "체육관": "체력단련장", "스포츠센터": "체력단련장", "휘트니스": "체력단련장", "크로스핏": "체력단련장",
        "당구": "당구장", "포켓볼": "당구장", "billiard": "당구장",
        "목욕탕": "목욕장업", "사우나": "목욕장업", "찜질방": "목욕장업", "목욕": "목욕장업",
        "썰매": "썰매장", "눈썰매": "썰매장",
        "골프": "골프연습장", "골프장": "골프연습장", "스크린골프": "골프연습장", "연습장": "골프연습장",
    }
    UPTAE_NAMES = set(SYNONYMS.values())

    # ── 전국 시/군/구 지명 사전 (약칭 → DB sigungu 값) ────────────
    KR_CITIES = {
        # 서울 자치구
        "강남": "강남구", "강동": "강동구", "강북": "강북구", "강서": "강서구",
        "관악": "관악구", "광진": "광진구", "구로": "구로구", "금천": "금천구",
        "노원": "노원구", "도봉": "도봉구", "동대문": "동대문구", "동작": "동작구",
        "마포": "마포구", "서대문": "서대문구", "서초": "서초구", "성동": "성동구",
        "성북": "성북구", "송파": "송파구", "양천": "양천구", "영등포": "영등포구",
        "용산": "용산구", "은평": "은평구", "종로": "종로구", "중랑": "중랑구",
        # 경기도
        "수원": "수원시", "성남": "성남시", "의정부": "의정부시", "안양": "안양시",
        "부천": "부천시", "광명": "광명시", "평택": "평택시", "안산": "안산시",
        "고양": "고양시", "과천": "과천시", "구리": "구리시", "남양주": "남양주시",
        "오산": "오산시", "시흥": "시흥시", "군포": "군포시", "하남": "하남시",
        "용인": "용인시", "파주": "파주시", "이천": "이천시", "안성": "안성시",
        "김포": "김포시", "화성": "화성시", "양주": "양주시", "포천": "포천시", "여주": "여주시",
        # 경상남도
        "창원": "창원시", "진주": "진주시", "통영": "통영시", "사천": "사천시",
        "김해": "김해시", "밀양": "밀양시", "거제": "거제시", "양산": "양산시",
        "함안": "함안군", "고성": "고성군", "남해": "남해군", "하동": "하동군",
        # 경상북도
        "포항": "포항시", "경주": "경주시", "김천": "김천시", "안동": "안동시",
        "구미": "구미시", "영주": "영주시", "영천": "영천시", "상주": "상주시",
        "문경": "문경시", "경산": "경산시",
        # 충청남도
        "천안": "천안시", "공주": "공주시", "보령": "보령시", "아산": "아산시",
        "서산": "서산시", "논산": "논산시", "계룡": "계룡시", "당진": "당진시",
        # 충청북도
        "청주": "청주시", "충주": "충주시", "제천": "제천시",
        # 전라남도
        "목포": "목포시", "여수": "여수시", "순천": "순천시", "나주": "나주시", "광양": "광양시",
        # 전라북도
        "전주": "전주시", "군산": "군산시", "익산": "익산시", "정읍": "정읍시",
        "남원": "남원시", "김제": "김제시",
        # 강원도
        "춘천": "춘천시", "원주": "원주시", "강릉": "강릉시", "동해": "동해시",
        "태백": "태백시", "속초": "속초시", "삼척": "삼척시",
        # 부산 자치구
        "해운대": "해운대구", "부산진": "부산진구", "동래": "동래구", "남구": "남구",
        "북구": "북구", "사하": "사하구", "금정": "금정구", "연제": "연제구",
        "수영": "수영구", "사상": "사상구", "기장": "기장군",
        # 인천 자치구
        "남동": "남동구", "부평": "부평구", "계양": "계양구", "서구": "서구",
        "연수": "연수구", "미추홀": "미추홀구", "중구": "중구", "강화": "강화군",
        # 대구 자치구
        "달서": "달서구", "달성": "달성군", "동구": "동구", "북구": "북구",
        "수성": "수성구", "중구": "중구",
    }

    # ── 멀티워드 쿼리 파싱 ─────────────────────────────────────────
    # "김해 헬스장" → synonym="체력단련장", sigungu_hint="김해시"
    q_parts = q.strip().split()
    synonym = None
    sigungu_hint = None
    text_parts = []

    for part in q_parts:
        if part in SYNONYMS:
            synonym = SYNONYMS[part]
        elif part in UPTAE_NAMES:
            synonym = part
        elif not sigungu and part in KR_CITIES:
            sigungu_hint = KR_CITIES[part]      # "김해" → "김해시"
        elif not sigungu and part.endswith(("시", "군", "구")) and len(part) >= 3:
            sigungu_hint = part                  # "김해시" → "김해시"
        else:
            text_parts.append(part)

    conditions = []

    # ── 검색어 조건 ──────────────────────────────────────────────
    if synonym:
        conditions.append(Business.uptae_nm == synonym)
    else:
        text_q = " ".join(text_parts) if text_parts else q
        conditions.append(
            or_(
                Business.bsn_nm.ilike(f"%{text_q}%"),
                Business.uptae_nm.ilike(f"%{text_q}%"),
            )
        )

    # 쿼리에서 추출한 시군구 힌트 적용
    if sigungu_hint and not sigungu:
        conditions.append(Business.sigungu.like(f"{sigungu_hint}%"))

    # ── 필터 조건 (인덱스 활용) ───────────────────────────────────
    # 시도 약칭 → 전체명 매핑 (DB에 "경상북도"로 저장, 필터는 "경북"으로 전송)
    SIDO_FULL = {
        "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시",
        "인천": "인천광역시", "광주": "광주광역시", "대전": "대전광역시",
        "울산": "울산광역시", "세종": "세종특별자치시", "경기": "경기도",
        "강원": "강원도", "충북": "충청북도", "충남": "충청남도",
        "전북": "전라북도", "전남": "전라남도", "경북": "경상북도",
        "경남": "경상남도", "제주": "제주특별자치도",
    }
    if sido and sido != "전국":
        full_sido = SIDO_FULL.get(sido, sido)
        # 약칭("경북") 또는 전체명("경상북도") 모두 매칭
        conditions.append(
            or_(Business.sido.like(f"{full_sido}%"), Business.sido.like(f"{sido}%"))
        )
    if sigungu:
        conditions.append(Business.sigungu.like(f"{sigungu}%"))
    if uptae:
        conditions.append(Business.uptae_nm == uptae)
    if status and status != "전체":
        if status == "영업중":
            conditions.append(or_(Business.status == "영업중", Business.status.like("%영업%")))
        else:
            conditions.append(Business.status == status)

    # ── COUNT (서브쿼리 없이 직접 카운트 → 빠름) ─────────────────
    count_stmt = select(func.count(Business.id))
    for cond in conditions:
        count_stmt = count_stmt.where(cond)
    total = (await db.execute(count_stmt)).scalar()

    # ── 데이터 조회 ───────────────────────────────────────────────
    stmt = select(Business)
    for cond in conditions:
        stmt = stmt.where(cond)
    stmt = stmt.order_by(Business.bsn_nm).offset((page - 1) * size).limit(size)
    rows = (await db.execute(stmt)).scalars().all()

    # 필터 없는 1페이지일 때만 상위 3건 FREE (지역/상태 필터 적용 시 FREE 없음)
    is_unfiltered = (not sido or sido == "전국") and (not status or status == "전체") and (not sigungu) and (not uptae) and page == 1

    items = []
    for i, b in enumerate(rows):
        rank = (page - 1) * size + i + 1
        revealed = str(b.id) in viewed_ids
        free = is_unfiltered and rank <= settings.FREE_VIEW_COUNT
        items.append(build_business_item(b, revealed, rank, free=free))

    log = SearchLog(
        user_id=current_user.id if current_user else None,
        query=q,
        result_cnt=total,
        ip=request.client.host if request and request.client else None,
    )
    db.add(log)
    await db.commit()

    return {"total": total, "page": page, "size": size, "items": items}


# ── 열람 ──────────────────────────────────────────────────────
@router.post("/reveal")
async def reveal_items(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ids = body.get("ids", [])
    if not ids:
        return {"message": "선택된 항목이 없습니다", "revealed": []}

    vl = await db.execute(
        select(ViewLog.business_id).where(
            ViewLog.user_id == current_user.id,
            ViewLog.business_id.in_(ids),
        )
    )
    already_viewed = {str(row[0]) for row in vl.fetchall()}
    new_ids = [i for i in ids if i not in already_viewed]

    if not new_ids:
        return {"message": "이미 열람한 항목입니다", "revealed": list(already_viewed)}

    # 가게별 포인트 계산 (전화번호 없으면 5P)
    new_bizs = (await db.execute(select(Business).where(Business.id.in_(new_ids)))).scalars().all()
    required_points = sum(get_point_cost(b) for b in new_bizs)

    if current_user.points < required_points:
        raise HTTPException(
            status_code=402,
            detail=f"포인트가 부족합니다. 필요: {required_points}P, 보유: {current_user.points}P",
        )

    current_user.points -= required_points
    db.add(PointLog(
        user_id=current_user.id,
        type="use",
        amount=-required_points,
        balance=current_user.points,
        description=f"{len(new_ids)}건 가게 정보 열람",
    ))

    for b in new_bizs:
        db.add(ViewLog(user_id=current_user.id, business_id=b.id, points_used=get_point_cost(b)))

    await db.commit()

    revealed_data = []
    for b in new_bizs:
        revealed_data.append({
            "id": str(b.id),
            "bsn_nm": b.bsn_nm,
            "uptae_nm": b.uptae_nm,
            "addr": b.addr,
            "road_addr": b.road_addr,
            "tel": b.tel,
            "status": normalize_status(b.status),
            "open_date": str(b.open_date) if b.open_date else None,
            "has_tel": bool(b.tel and b.tel.strip()),
        })

    return {
        "message": f"{len(new_ids)}건 열람 완료",
        "points_used": required_points,
        "remaining_points": current_user.points,
        "revealed": revealed_data,
    }


# ── 엑셀 다운로드 ─────────────────────────────────────────────
@router.post("/excel")
async def download_excel(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="선택된 항목이 없습니다")

    vl_res = await db.execute(
        select(ViewLog.business_id).where(
            ViewLog.user_id == current_user.id,
            ViewLog.business_id.in_(ids),
        )
    )
    already_viewed = {str(row[0]) for row in vl_res.fetchall()}
    new_ids = [i for i in ids if i not in already_viewed]

    if new_ids:
        new_bizs = (await db.execute(select(Business).where(Business.id.in_(new_ids)))).scalars().all()
        required_points = sum(get_point_cost(b) for b in new_bizs)

        if current_user.points < required_points:
            raise HTTPException(
                status_code=402,
                detail=f"포인트 부족. 필요: {required_points}P, 보유: {current_user.points}P",
            )
        current_user.points -= required_points
        db.add(PointLog(
            user_id=current_user.id,
            type="use",
            amount=-required_points,
            balance=current_user.points,
            description=f"{len(new_ids)}건 엑셀 다운로드",
        ))
        for b in new_bizs:
            db.add(ViewLog(user_id=current_user.id, business_id=b.id, points_used=get_point_cost(b)))
        await db.commit()

    all_ids = list(set(ids))
    result = await db.execute(select(Business).where(Business.id.in_(all_ids)))
    businesses = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "가게정보"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["순번", "상호명", "업종", "영업상태", "도로명주소", "주소", "전화번호", "개업일자", "시도", "시군구"]
    col_widths = [6, 25, 15, 10, 40, 40, 16, 12, 8, 10]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, b in enumerate(businesses, 2):
        row_data = [
            row_idx - 1,
            b.bsn_nm,
            b.uptae_nm or "",
            normalize_status(b.status),
            b.road_addr or "",
            b.addr or "",
            b.tel or "",
            str(b.open_date) if b.open_date else "",
            b.sido or "",
            b.sigungu or "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if col_idx == 1:
                cell.alignment = center

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = quote("AIDB_가게정보.xlsx")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
